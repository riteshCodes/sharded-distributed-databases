from os import path
from pathlib import Path

import matplotlib.pyplot as plt
from matplotlib import ticker
from matplotlib.ticker import FormatStrFormatter
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Path to profiling data
FILE_PATH = path.join(path.dirname(Path(__file__)), Path('profiling/client_profiling.xlsx'))
# Set display options pandas
pd.options.display.float_format = '{:.10E}'.format


def get_data_from_excel(*, with_sheet_name: str, header_value=5,
                        with_columns=None):
    if with_columns is None:
        with_columns = ['Execution_Time (seconds)', 'Response_Time (seconds)']
    data_frame = pd.read_excel(FILE_PATH, header=header_value, sheet_name=with_sheet_name, usecols=with_columns)
    return data_frame.dropna()


def visualize_data(*, function, sheet_name, with_columns):
    # Data configurations
    default_x_ticks = np.arange(4)
    keys_range = np.power(10, default_x_ticks)

    # Store latency data in a Pandas DataFrame
    # data = {'keys': keys_range,'latency': measurements}
    # df = pd.DataFrame(data)

    df = get_data_from_excel(with_sheet_name=sheet_name, with_columns=with_columns)

    # Plot configurations
    fig, axs = plt.subplots()
    for col in df.columns:
        plt.plot(default_x_ticks, df[col], marker='*', label=function)

    plt.xticks(default_x_ticks, keys_range)
    x_label = 'Number of key-value pairs'  # default x label
    function_label = function.split()[0].lower()
    if function_label == 'ping':
        x_label = 'Number of ping requests (string messages)'
    plt.xlabel(x_label)

    formatter = ticker.ScalarFormatter(useMathText=True)  # scientific notation
    formatter.set_scientific(True)
    formatter.set_powerlimits((0,0))
    axs.yaxis.set_major_formatter(formatter)
    plt.ylabel('Average Response Time (seconds)')  # milliseconds(ms), 1 second = 1000 milliseconds

    # plt.title(f'Total Response Time Per Given Key-Value Pairs')
    plt.legend(framealpha=1)
    plt.grid(True)
    plt.savefig(f'assets/{function.split()[0].lower()}.pdf', dpi=1200)
    plt.savefig(f'assets/{function.split()[0].lower()}.svg', dpi=1200)
    plt.show()


def visualize_data_combined_operations(*, with_columns):
    # Data configurations
    default_x_ticks = np.arange(4)
    keys_range = np.power(10, default_x_ticks)

    # Store latency data in a Pandas DataFrame
    # data = {'keys': keys_range,'latency': measurements}
    # df = pd.DataFrame(data)

    # Plot configurations
    fig, axs = plt.subplots()

    profiling_file = pd.ExcelFile(FILE_PATH)

    # Iterate through each worksheet
    for sheet in profiling_file.sheet_names:
        if sheet == 'Ping Message':
            continue
        df = get_data_from_excel(with_sheet_name=sheet, with_columns=with_columns)
        for col in df.columns:
            plt.plot(default_x_ticks, df[col], marker='*', label=sheet)

    plt.xticks(default_x_ticks, keys_range)
    plt.xlabel('Number of key-value pairs')

    formatter = ticker.ScalarFormatter(useMathText=True)  # scientific notation
    formatter.set_scientific(True)
    formatter.set_powerlimits((-1, 4))
    axs.yaxis.set_major_formatter(formatter)
    plt.ylabel('Average Response Time (seconds)')  # milliseconds(ms), 1 second = 1000 milliseconds

    # plt.title(f'Average Response Time Per Given Key-Value Pairs (With Middleware)')
    plt.legend(framealpha=1)
    plt.grid(True)
    plt.savefig(f'assets/combined_visualization.pdf', dpi=1200)
    plt.show()


def write_to_excel_wall_time(*, sheet_name: str, start_row: int = 7, data):
    df = pd.DataFrame({'Response_Time (seconds)': data})
    wb = load_workbook(FILE_PATH)
    ws = wb[sheet_name]

    for index, row in df.iterrows():
        cell = 'C%d' % (index + start_row)
        ws[cell] = row[0]

    wb.save(FILE_PATH)


def write_to_excel_exec_time(*, sheet_name: str, start_row: int = 7, data):
    df = pd.DataFrame({'Execution_Time (seconds)': data})
    wb = load_workbook(FILE_PATH)
    ws = wb[sheet_name]

    for index, row in df.iterrows():
        cell = 'B%d' % (index + start_row)
        ws[cell] = row[0]

    wb.save(FILE_PATH)


if __name__ == '__main__':
    visualize_data_combined_operations(with_columns=['Response_Time (seconds)'])

    # Visualization from client side to middleware
    """
    visualize_data(function='Ping (String) Request', sheet_name='Ping Message',
                   with_columns=['Response_Time (seconds)'])
    """
    # Operations visualization
    visualize_data(function='Set Function', sheet_name='Set Function',
                   with_columns=['Response_Time (seconds)'])

    visualize_data(function='Get Function', sheet_name='Get Function',
                   with_columns=['Response_Time (seconds)'])
    visualize_data(function='Get-Range Function', sheet_name='Get-Range Function',
                   with_columns=['Response_Time (seconds)'])
    visualize_data(function='Delete Function', sheet_name='Delete Function',
                   with_columns=['Response_Time (seconds)'])
